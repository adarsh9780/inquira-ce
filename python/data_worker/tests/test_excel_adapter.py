from __future__ import annotations

import datetime as dt
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook, load_workbook

from inquira_data_worker.adapters.registry import get_adapter
from inquira_data_worker.errors import AdapterError
from inquira_data_worker.models import AdapterRequest, MaterializeRequest


def write_workbook(path: Path) -> None:
    workbook = Workbook()
    sales = workbook.active
    sales.title = "Sales 2026"
    sales.append(["id", "customer", "amount", "ordered_on", "active"])
    sales.append([1, "José", 12.5, dt.date(2026, 7, 1), True])
    sales.append([2, "東京", None, dt.date(2026, 7, 2), False])

    notes = workbook.create_sheet("Internal notes")
    notes.sheet_state = "hidden"
    notes.append(["note"])
    notes.append(["private"])

    workbook.create_sheet("Empty")
    workbook.save(path)
    workbook.close()


def sheet_id(name: str) -> str:
    return f"sheet:{name}"


def test_excel_discovery_lists_every_sheet_with_visibility_dimensions_and_schema(tmp_path: Path) -> None:
    path = tmp_path / "Quarterly ünicode.xlsx"
    write_workbook(path)

    result = get_adapter("EXCEL").discover(AdapterRequest(source_path=str(path)))

    assert result.adapter_kind == "excel"
    assert result.source_path == str(path.resolve())
    assert [item.id for item in result.objects] == [
        sheet_id("Sales 2026"),
        sheet_id("Internal notes"),
        sheet_id("Empty"),
    ]
    sales, notes, empty = result.objects
    assert sales.name == "Sales 2026"
    assert sales.kind == "sheet"
    assert sales.metadata == {"visibility": "visible", "row_count": 2, "column_count": 5, "selectable": True}
    assert [(column.name, column.data_type) for column in sales.columns] == [
        ("id", "BIGINT"),
        ("customer", "VARCHAR"),
        ("amount", "DOUBLE"),
        ("ordered_on", "DATE"),
        ("active", "BOOLEAN"),
    ]
    assert notes.metadata["visibility"] == "hidden"
    assert notes.metadata["selectable"] is True
    assert empty.columns == []
    assert empty.metadata == {"visibility": "visible", "row_count": 0, "column_count": 0, "selectable": False}
    assert result.fingerprint.startswith("sha256:")


def test_excel_preview_requires_a_sheet_and_handles_blank_duplicate_headers(tmp_path: Path) -> None:
    path = tmp_path / "headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Messy"
    sheet.append(["name", "", "name", None])
    sheet.append(["Ada", 10, "Lovelace", None])
    sheet.append(["東京", None, "Japan", "extra"])
    workbook.save(path)
    workbook.close()
    adapter = get_adapter("excel")

    with pytest.raises(AdapterError, match="sheet"):
        adapter.preview(AdapterRequest(source_path=str(path)), limit=10)

    preview = adapter.preview(
        AdapterRequest(source_path=str(path), options={"source_object_id": sheet_id("Messy")}),
        limit=1,
    )
    assert [column.name for column in preview.columns] == ["name", "column_2", "name_2", "column_4"]
    assert preview.rows == [{"name": "Ada", "column_2": 10, "name_2": "Lovelace", "column_4": None}]
    assert preview.truncated is True


def test_excel_preview_preserves_formula_text_or_uses_cached_values(tmp_path: Path) -> None:
    path = tmp_path / "formulas.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calculated"
    sheet.append(["amount", "doubled"])
    sheet.append([4, "=A2*2"])
    workbook.save(path)
    workbook.close()
    adapter = get_adapter("excel")
    source_object_id = sheet_id("Calculated")

    formula_preview = adapter.preview(
        AdapterRequest(source_path=str(path), options={"source_object_id": source_object_id, "formula_mode": "formula"}),
        limit=10,
    )
    assert formula_preview.rows[0]["doubled"] == "=A2*2"

    cached_preview = adapter.preview(
        AdapterRequest(source_path=str(path), options={"source_object_id": source_object_id}),
        limit=10,
    )
    assert cached_preview.rows[0]["doubled"] is None

    with pytest.raises(AdapterError, match="formula_mode"):
        adapter.preview(
            AdapterRequest(source_path=str(path), options={"source_object_id": source_object_id, "formula_mode": "calculate"}),
            limit=10,
        )


@pytest.mark.parametrize("limit", [0, -1, 1001])
def test_excel_preview_rejects_invalid_limits(limit: int, tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    write_workbook(path)
    with pytest.raises(AdapterError, match="limit"):
        get_adapter("excel").preview(
            AdapterRequest(source_path=str(path), options={"source_object_id": sheet_id("Sales 2026")}),
            limit=limit,
        )


def test_excel_materializes_multiple_selected_sheets_as_separate_parquet_outputs(tmp_path: Path) -> None:
    path = tmp_path / "multiple.xlsx"
    write_workbook(path)
    target = tmp_path / "snapshot"

    result = get_adapter("excel").materialize(MaterializeRequest(
        source_path=str(path),
        target_dir=str(target),
        selected_object_ids=[sheet_id("Sales 2026"), sheet_id("Internal notes")],
    ))

    assert [output.source_object_id for output in result.outputs] == [
        sheet_id("Sales 2026"),
        sheet_id("Internal notes"),
    ]
    assert [output.name for output in result.outputs] == ["Sales 2026", "Internal notes"]
    assert len({output.relative_path for output in result.outputs}) == 2
    assert all(".." not in output.relative_path and output.relative_path.endswith(".parquet") for output in result.outputs)
    assert [output.row_count for output in result.outputs] == [2, 1]
    for output in result.outputs:
        output_path = target / output.relative_path
        assert output.byte_size == output_path.stat().st_size
        assert duckdb.sql(f"SELECT count(*) FROM read_parquet('{output_path}')").fetchone()[0] == output.row_count


def test_excel_materialization_has_no_preview_row_limit_and_preserves_mixed_values_safely(tmp_path: Path) -> None:
    path = tmp_path / "large.xlsx"
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Rows")
    sheet.append(["id", "mixed", "occurred_at"])
    for index in range(257):
        mixed = index if index % 2 == 0 else f"row-{index}"
        sheet.append([index, mixed, dt.datetime(2026, 7, 1, 12, index % 60)])
    workbook.save(path)
    workbook.close()

    target = tmp_path / "large-output"
    result = get_adapter("excel").materialize(MaterializeRequest(
        source_path=str(path),
        target_dir=str(target),
        selected_object_ids=[sheet_id("Rows")],
    ))

    output = result.outputs[0]
    assert output.row_count == 257
    assert [(column.name, column.data_type) for column in output.columns] == [
        ("id", "BIGINT"),
        ("mixed", "VARCHAR"),
        ("occurred_at", "TIMESTAMP"),
    ]
    assert duckdb.sql(f"SELECT mixed FROM read_parquet('{target / output.relative_path}') ORDER BY id LIMIT 2").fetchall() == [
        ("0",),
        ("row-1",),
    ]


def test_excel_rejects_missing_duplicate_unknown_and_empty_sheet_selections(tmp_path: Path) -> None:
    path = tmp_path / "selection.xlsx"
    write_workbook(path)
    adapter = get_adapter("excel")
    cases = [
        ([], "sheet"),
        ([sheet_id("Sales 2026"), sheet_id("Sales 2026")], "unique"),
        ([sheet_id("Renamed")], "no longer exists"),
        ([sheet_id("Empty")], "empty"),
    ]
    for index, (selected, message) in enumerate(cases):
        with pytest.raises(AdapterError, match=message):
            adapter.materialize(MaterializeRequest(
                source_path=str(path),
                target_dir=str(tmp_path / f"target-{index}"),
                selected_object_ids=selected,
            ))


def test_excel_rejects_legacy_empty_corrupt_and_wrong_extension_sources(tmp_path: Path) -> None:
    adapter = get_adapter("excel")
    files = [
        ("legacy.xls", b"legacy", r"\.xlsx"),
        ("empty.xlsx", b"", "empty"),
        ("corrupt.xlsx", b"not an office zip", "read"),
    ]
    for filename, content, message in files:
        path = tmp_path / filename
        path.write_bytes(content)
        with pytest.raises(AdapterError, match=message):
            adapter.discover(AdapterRequest(source_path=str(path)))


def test_excel_fingerprint_changes_when_any_sheet_changes(tmp_path: Path) -> None:
    path = tmp_path / "refresh.xlsx"
    write_workbook(path)
    adapter = get_adapter("excel")
    first = adapter.discover(AdapterRequest(source_path=str(path))).fingerprint
    assert adapter.discover(AdapterRequest(source_path=str(path))).fingerprint == first

    workbook = load_workbook(path)
    workbook["Sales 2026"].append([3, "Ada", 20, dt.date(2026, 7, 3), True])
    workbook.save(path)
    workbook.close()
    assert adapter.discover(AdapterRequest(source_path=str(path))).fingerprint != first


def test_excel_materialization_rejects_a_workbook_changed_during_inspection(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    import inquira_data_worker.adapters.excel as excel_module

    path = tmp_path / "changing.xlsx"
    write_workbook(path)
    original = excel_module._analyse_sheet
    changed = False

    def analyse_and_change(sheet):
        nonlocal changed
        result = original(sheet)
        if not changed:
            with path.open("ab") as handle:
                handle.write(b"changed-during-inspection")
            changed = True
        return result

    monkeypatch.setattr(excel_module, "_analyse_sheet", analyse_and_change)
    with pytest.raises(AdapterError, match="changed"):
        get_adapter("excel").materialize(MaterializeRequest(
            source_path=str(path),
            target_dir=str(tmp_path / "changed-output"),
            selected_object_ids=[sheet_id("Sales 2026")],
        ))

"""Streaming adapter for modern Excel workbooks."""

from __future__ import annotations

import datetime as dt
import hashlib
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import duckdb
from openpyxl import load_workbook

from ..errors import AdapterError
from ..models import (
    AdapterRequest,
    Column,
    Discovery,
    Materialization,
    MaterializedOutput,
    MaterializeRequest,
    Preview,
    SourceObject,
)
from .file import MAX_PREVIEW_ROWS, _fingerprint, _json_value, _sql_string

SHEET_PREFIX = "sheet:"
INSERT_BATCH_SIZE = 1000


@dataclass(frozen=True)
class _SheetAnalysis:
    name: str
    visibility: str
    columns: list[Column]
    row_count: int


def _sheet_id(name: str) -> str:
    return SHEET_PREFIX + name


def _sheet_name(object_id: str) -> str:
    value = str(object_id or "")
    if not value.startswith(SHEET_PREFIX) or not value[len(SHEET_PREFIX):]:
        raise AdapterError("invalid_selection", "Select a valid workbook sheet.")
    return value[len(SHEET_PREFIX):]


def _trim_row(row: Iterable[Any]) -> list[Any]:
    values = list(row)
    while values and values[-1] is None:
        values.pop()
    return values


def _nonempty_rows(sheet: Any) -> Iterable[list[Any]]:
    for row in sheet.iter_rows(values_only=True):
        values = _trim_row(row)
        if values:
            yield values


def _headers(values: list[Any], width: int) -> list[str]:
    result: list[str] = []
    used: dict[str, int] = {}
    for index in range(width):
        raw = values[index] if index < len(values) else None
        base = str(raw).strip() if raw is not None else ""
        if not base:
            base = f"column_{index + 1}"
        key = base.casefold()
        count = used.get(key, 0) + 1
        used[key] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def _value_type(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, dt.datetime):
        return "DATE" if value.time() == dt.time() else "TIMESTAMP"
    if isinstance(value, dt.date):
        return "DATE"
    if isinstance(value, dt.time):
        return "TIME"
    if isinstance(value, int) and -(2**63) <= value < 2**63:
        return "BIGINT"
    if isinstance(value, float):
        return "DOUBLE"
    return "VARCHAR"


def _merge_type(current: str | None, incoming: str | None) -> str | None:
    if incoming is None:
        return current
    if current is None or current == incoming:
        return incoming
    if {current, incoming} <= {"BIGINT", "DOUBLE"}:
        return "DOUBLE"
    if {current, incoming} <= {"DATE", "TIMESTAMP"}:
        return "TIMESTAMP"
    return "VARCHAR"


def _analyse_sheet(sheet: Any) -> _SheetAnalysis:
    iterator = iter(_nonempty_rows(sheet))
    header_values = next(iterator, [])
    width = len(header_values)
    inferred: list[str | None] = [None] * width
    row_count = 0
    for row in iterator:
        width = max(width, len(row))
        if len(inferred) < width:
            inferred.extend([None] * (width - len(inferred)))
        for index in range(width):
            value = row[index] if index < len(row) else None
            inferred[index] = _merge_type(inferred[index], _value_type(value))
        row_count += 1
    names = _headers(header_values, width)
    columns = [Column(name=name, data_type=inferred[index] or "VARCHAR") for index, name in enumerate(names)]
    return _SheetAnalysis(
        name=sheet.title,
        visibility=str(sheet.sheet_state or "visible"),
        columns=columns,
        row_count=row_count,
    )


def _quote_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _convert(value: Any, data_type: str) -> Any:
    if value is None:
        return None
    if data_type == "VARCHAR":
        normalized = _json_value(value)
        return normalized if isinstance(normalized, str) else str(normalized)
    if data_type == "DOUBLE" and isinstance(value, int):
        return float(value)
    if data_type == "DATE" and isinstance(value, dt.datetime):
        return value.date()
    if data_type == "TIMESTAMP" and isinstance(value, dt.date) and not isinstance(value, dt.datetime):
        return dt.datetime.combine(value, dt.time())
    return value


class ExcelAdapter:
    kind = "excel"
    suffix = ".xlsx"

    def _source(self, value: str) -> Path:
        if not str(value or "").strip():
            raise AdapterError("invalid_params", "A source path is required.")
        raw = Path(value).expanduser()
        if not raw.exists():
            raise AdapterError("source_not_found", f"Source file does not exist: {raw}")
        if not raw.is_file():
            raise AdapterError("source_not_file", "Source path must be a regular file.")
        if raw.suffix.lower() != self.suffix:
            raise AdapterError("source_extension_mismatch", "Expected a .xlsx file extension.")
        if raw.stat().st_size == 0:
            raise AdapterError("source_unreadable", "Could not read excel source: file is empty.")
        return raw.resolve(strict=True)

    def _formula_mode(self, options: dict[str, Any]) -> str:
        mode = str(options.get("formula_mode") or "cached").strip().lower()
        if mode not in {"cached", "formula"}:
            raise AdapterError("invalid_option", "formula_mode must be cached or formula.")
        return mode

    def _open(self, path: Path, formula_mode: str):
        try:
            return load_workbook(
                path,
                read_only=True,
                data_only=formula_mode == "cached",
                keep_links=False,
            )
        except Exception as exc:
            raise AdapterError("source_unreadable", f"Could not read excel source: {exc}") from exc

    def _selected_sheet_id(self, request: AdapterRequest) -> str:
        selected = request.source_object_id or str(request.options.get("source_object_id") or "")
        if not selected:
            raise AdapterError("source_selection_required", "Select a workbook sheet to preview.")
        return selected

    def discover(self, request: AdapterRequest) -> Discovery:
        path = self._source(request.source_path)
        formula_mode = self._formula_mode(request.options)
        before = _fingerprint(path)
        workbook = self._open(path, formula_mode)
        try:
            analyses = [_analyse_sheet(sheet) for sheet in workbook.worksheets]
        finally:
            workbook.close()
        after = _fingerprint(path)
        if after != before:
            raise AdapterError("source_changed", "Source file changed during discovery; inspect it again.")
        return Discovery(
            adapter_kind=self.kind,
            source_path=str(path),
            fingerprint=after,
            objects=[SourceObject(
                id=_sheet_id(item.name),
                name=item.name,
                kind="sheet",
                columns=item.columns,
                metadata={
                    "visibility": item.visibility,
                    "row_count": item.row_count,
                    "column_count": len(item.columns),
                    "selectable": bool(item.columns),
                },
            ) for item in analyses],
        )

    def preview(self, request: AdapterRequest, limit: int) -> Preview:
        if limit < 1 or limit > MAX_PREVIEW_ROWS:
            raise AdapterError("invalid_preview_limit", f"Preview limit must be between 1 and {MAX_PREVIEW_ROWS}.")
        path = self._source(request.source_path)
        formula_mode = self._formula_mode(request.options)
        name = _sheet_name(self._selected_sheet_id(request))
        workbook = self._open(path, formula_mode)
        try:
            if name not in workbook.sheetnames:
                raise AdapterError("source_selection_missing", f"Selected sheet {name} no longer exists.")
            analysis = _analyse_sheet(workbook[name])
        finally:
            workbook.close()
        if not analysis.columns:
            return Preview(columns=[], rows=[], truncated=False)

        workbook = self._open(path, formula_mode)
        try:
            iterator = iter(_nonempty_rows(workbook[name]))
            next(iterator, None)
            rows: list[dict[str, Any]] = []
            for values in iterator:
                row = {
                    column.name: _json_value(_convert(values[index], column.data_type)) if index < len(values) else None
                    for index, column in enumerate(analysis.columns)
                }
                rows.append(row)
                if len(rows) > limit:
                    break
        finally:
            workbook.close()
        return Preview(columns=analysis.columns, rows=rows[:limit], truncated=len(rows) > limit)

    def materialize(self, request: MaterializeRequest) -> Materialization:
        selected = request.selected_object_ids
        if not selected:
            raise AdapterError("source_selection_required", "Select at least one workbook sheet.")
        if len(set(selected)) != len(selected):
            raise AdapterError("invalid_selection", "Selected workbook sheets must be unique.")
        path = self._source(request.source_path)
        formula_mode = self._formula_mode(request.options)
        target = Path(request.target_dir).expanduser().resolve()
        if target.exists() and (not target.is_dir() or any(target.iterdir())):
            raise AdapterError("target_not_empty", "Materialization target must be empty.")

        before = _fingerprint(path)
        names = [_sheet_name(item) for item in selected]
        workbook = self._open(path, formula_mode)
        try:
            missing = [name for name in names if name not in workbook.sheetnames]
            if missing:
                raise AdapterError("source_selection_missing", f"Selected sheet {missing[0]} no longer exists.")
            analyses = {name: _analyse_sheet(workbook[name]) for name in names}
        finally:
            workbook.close()
        empty = [name for name in names if not analyses[name].columns]
        if empty:
            raise AdapterError("empty_sheet", f"Selected sheet {empty[0]} is empty.")

        target.mkdir(parents=True, exist_ok=True)
        workbook = self._open(path, formula_mode)
        outputs: list[MaterializedOutput] = []
        try:
            for index, (object_id, name) in enumerate(zip(selected, names, strict=True)):
                analysis = analyses[name]
                digest = hashlib.sha256(object_id.encode("utf-8")).hexdigest()[:12]
                relative_path = f"sheet-{index + 1}-{digest}.parquet"
                output_path = target / relative_path
                self._write_sheet(workbook[name], analysis, output_path)
                outputs.append(MaterializedOutput(
                    source_object_id=object_id,
                    name=name,
                    relative_path=relative_path,
                    format="parquet",
                    columns=analysis.columns,
                    row_count=analysis.row_count,
                    byte_size=output_path.stat().st_size,
                ))
        except AdapterError:
            shutil.rmtree(target, ignore_errors=True)
            raise
        except Exception as exc:
            shutil.rmtree(target, ignore_errors=True)
            raise AdapterError("materialization_failed", f"Could not materialize excel source: {exc}") from exc
        finally:
            workbook.close()

        after = _fingerprint(path)
        if after != before:
            shutil.rmtree(target, ignore_errors=True)
            raise AdapterError("source_changed", "Source file changed during materialization; refresh again.")
        return Materialization(fingerprint=after, outputs=outputs)

    def _write_sheet(self, sheet: Any, analysis: _SheetAnalysis, output: Path) -> None:
        connection = duckdb.connect()
        try:
            definitions = ", ".join(
                f"{_quote_identifier(column.name)} {column.data_type}" for column in analysis.columns
            )
            connection.execute(f"CREATE TABLE sheet_data ({definitions})")
            placeholders = ", ".join("?" for _ in analysis.columns)
            iterator = iter(_nonempty_rows(sheet))
            next(iterator, None)
            batch: list[list[Any]] = []
            for values in iterator:
                batch.append([
                    _convert(values[index] if index < len(values) else None, column.data_type)
                    for index, column in enumerate(analysis.columns)
                ])
                if len(batch) == INSERT_BATCH_SIZE:
                    connection.executemany(f"INSERT INTO sheet_data VALUES ({placeholders})", batch)
                    batch.clear()
            if batch:
                connection.executemany(f"INSERT INTO sheet_data VALUES ({placeholders})", batch)
            connection.execute(f"COPY sheet_data TO {_sql_string(output)} (FORMAT PARQUET)")
        finally:
            connection.close()
